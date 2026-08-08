from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from core.models import Relatorio
from .services import AnalyticsService
from django.db.models import Count, Sum
from users.permissions import IsVendedor, IsSupervisorOrAdmin, IsAdmin, IsAdminOrAdminCliente
from django.http import HttpResponse
from django.utils.text import slugify
class MeuDesempenhoView(APIView):
    """
    Retorna o desempenho individual do vendedor logado.
    """
    permission_classes = [IsAuthenticated, IsVendedor]

    def get(self, request):
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')

        queryset = Relatorio.objects.filter(vendedor=request.user)

        queryset_filtrado = AnalyticsService.filter_by_date(
            queryset=queryset, 
            data_inicio=data_inicio, 
            data_fim=data_fim
        )

        dados_dashboard = AnalyticsService.get_dashboard_data(queryset_filtrado)
        return Response(dados_dashboard)

class LojaDesempenhoView(APIView):
    """
    Retorna o desempenho consolidado da loja do usuario logado.
    ADMIN_CLIENTE pode filtrar por qualquer loja do seu cliente via ?loja_id=.
    """
    permission_classes = [IsAuthenticated, IsSupervisorOrAdmin]

    def get(self, request):
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')
        loja_id = request.query_params.get('loja_id')

        # Determina a loja alvo
        loja_alvo = None

        if loja_id:
            # ADMIN/ADMIN_CLIENTE filtrando por loja especifica
            try:
                from core.models import Loja
                loja_alvo = Loja.objects.get(id=loja_id)
            except Loja.DoesNotExist:
                return Response({"detail": "Loja nao encontrada."}, status=404)

            # ADMIN_CLIENTE: validar que a loja pertence ao seu cliente
            if request.user.cargo == 'ADMIN_CLIENTE':
                if loja_alvo.cliente != request.user.cliente:
                    return Response(
                        {"detail": "Loja nao pertence ao seu cliente."},
                        status=403,
                    )
        elif request.user.loja:
            loja_alvo = request.user.loja

        if not loja_alvo:
            return Response(
                {"detail": "Usuario nao possui uma loja vinculada. Use o parametro ?loja_id= para selecionar uma loja."},
                status=400,
            )

        queryset = Relatorio.objects.filter(vendedor__loja=loja_alvo)

        queryset_filtrado = AnalyticsService.filter_by_date(
            queryset=queryset, 
            data_inicio=data_inicio, 
            data_fim=data_fim
        )

        dados_dashboard = AnalyticsService.get_dashboard_data(queryset_filtrado)

        ranking = (
            queryset_filtrado.filter(venda_fechada=True)
            .values('vendedor__first_name', 'vendedor__last_name')
            .annotate(
                total_vendas=Count('id'),
                valor_total=Sum('valor_venda')
            )
            .order_by('-valor_total')
        )
        
        dados_dashboard['ranking_vendedores'] = list(ranking)
        return Response(dados_dashboard)

class VisaoGeralView(APIView):
    """
    Retorna o desempenho global do sistema com comparativo entre lojas.
    ADMIN: ve todos os clientes. ADMIN_CLIENTE: ve apenas seu cliente.
    Pode ser filtrado por uma loja especifica via query_params.
    """
    permission_classes = [IsAuthenticated, IsAdminOrAdminCliente]

    def get(self, request):
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')
        loja_id = request.query_params.get('loja_id')

        # Escopo base: ADMIN ve tudo, ADMIN_CLIENTE ve so seu cliente
        queryset = Relatorio.objects.all()

        if request.user.cargo == 'ADMIN_CLIENTE' and request.user.cliente:
            queryset = queryset.filter(vendedor__loja__cliente=request.user.cliente)

        # SE houver loja_id, filtramos o queryset ANTES de mandar para o Service
        if loja_id:
            queryset = queryset.filter(vendedor__loja_id=loja_id)

        # Filtro de datas via Service
        queryset_filtrado = AnalyticsService.filter_by_date(
            queryset=queryset, 
            data_inicio=data_inicio, 
            data_fim=data_fim
        )

        # Gera os dados base (KPIs, Gráficos, Tabela) já filtrados
        dados_dashboard = AnalyticsService.get_dashboard_data(queryset_filtrado)

        # Métrica Extra: Comparativo entre Lojas
        comparativo = (
            queryset_filtrado.filter(venda_fechada=True)
            .values('vendedor__loja__nome')
            .annotate(
                quantidade_vendas=Count('id'),
                valor_total=Sum('valor_venda')
            )
            .order_by('-valor_total')
        )
        
        comparativo_formatado = [
            {
                "loja": item['vendedor__loja__nome'] or "Sem Loja Vinculada",
                "quantidade_vendas": item['quantidade_vendas'],
                "valor_total": item['valor_total']
            }
            for item in comparativo
        ]

        dados_dashboard['comparativo_lojas'] = comparativo_formatado

        return Response(dados_dashboard)
    

#==========================#========================== analise essa parte para baixo ==========================#==========================
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
# Presumo que o Relatorio já está importado lá em cima: from core.models import Relatorio

class ExportarDadosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, formato, *args, **kwargs):
        # 1. Capturar os parâmetros obrigatórios
        loja_id = request.query_params.get('loja_id')
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')

        if not loja_id or not data_inicio or not data_fim:
            return HttpResponse("Loja e periodo sao obrigatorios.", status=400)

        # 2. Validar que a loja pertence ao cliente do usuario
        if request.user.cargo == 'ADMIN_CLIENTE' and request.user.cliente:
            from core.models import Loja
            try:
                loja_obj = Loja.objects.get(id=loja_id)
            except Loja.DoesNotExist:
                return HttpResponse("Loja nao encontrada.", status=404)
            if loja_obj.cliente != request.user.cliente:
                return HttpResponse("Loja nao pertence ao seu cliente.", status=403)

        # 3. Filtrar os dados no Banco de Dados
        queryset = Relatorio.objects.filter(
            vendedor__loja_id=loja_id,
            data_hora__date__gte=parse_date(data_inicio),
            data_hora__date__lte=parse_date(data_fim)
        ).order_by('data_hora')

        # 3. DEPOIS: Descobrir o nome da loja a partir do queryset para usar no nome do arquivo
        primeiro_registro = queryset.first()
        if primeiro_registro and primeiro_registro.vendedor and primeiro_registro.vendedor.loja:
            nome_loja_slug = slugify(primeiro_registro.vendedor.loja.nome).replace('-', '_')
        else:
            nome_loja_slug = f"loja_{loja_id}"

        # 4. Definir os cabeçalhos das colunas
        cabecalhos = [
            "Horário", 
            "Colaborador", 
            "Loja", 
            "Cliente", 
            "Status", 
            "Valor Venda (R$)", 
            "Motivo / Métrica", 
            "Observações"
        ]

        # === EXPORTAÇÃO PARA CSV ===
        if formato.lower() == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="relatorio_{nome_loja_slug}_{data_inicio}_a_{data_fim}.csv"'
            
            # Escrever o BOM para o Excel abrir o CSV com acentos corretos em PT-BR
            response.write('\ufeff'.encode('utf8'))
            
            writer = csv.writer(response, delimiter=';')
            writer.writerow(cabecalhos)

            for item in queryset:
                status = "Concretizada" if item.venda_fechada else "Não Concretizada"
                valor = f"{item.valor_venda:.2f}" if item.venda_fechada and item.valor_venda else "-"
                vendedor_nome = f"{item.vendedor.first_name} {item.vendedor.last_name}" if item.vendedor else "N/A"
                loja_nome = item.vendedor.loja.nome if item.vendedor and item.vendedor.loja else "Sem loja"
                motivo = item.metrica.nome if not item.venda_fechada and item.metrica else "-"
                
                writer.writerow([
                    item.data_hora.strftime('%d/%m/%Y %H:%M') if item.data_hora else "-",
                    vendedor_nome,
                    loja_nome,
                    item.cliente_nome or "Não informado",
                    status,
                    valor,
                    motivo,
                    item.observacoes or ""
                ])
            return response

        # === EXPORTAÇÃO PARA EXCEL (XLSX) ===
        elif formato.lower() == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório de Vendas"

            # Aplicar estilos estéticos para o cabeçalho
            fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            preenchimento_cabecalho = PatternFill(start_color="4D7BAB", end_color="4D7BAB", fill_type="solid")
            alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border_fina = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )

            # Escrever cabeçalhos estilizados
            ws.append(cabecalhos)
            for cell in ws[1]:
                cell.font = fonte_cabecalho
                cell.fill = preenchimento_cabecalho
                cell.alignment = alinhamento_centro
                cell.border = border_fina
            
            ws.row_dimensions[1].height = 28

            # Inserir as linhas de dados
            font_dados = Font(name="Arial", size=10)
            for item in queryset:
                status = "Concretizada" if item.venda_fechada else "Não Concretizada"
                valor = float(item.valor_venda) if item.venda_fechada and item.valor_venda else 0.0
                vendedor_nome = f"{item.vendedor.first_name} {item.vendedor.last_name}" if item.vendedor else "N/A"
                loja_nome = item.vendedor.loja.nome if item.vendedor and item.vendedor.loja else "Sem loja"
                motivo = item.metrica.nome if not item.venda_fechada and item.metrica else "-"
                
                linha = [
                    item.data_hora.strftime('%d/%m/%Y %H:%M') if item.data_hora else "-",
                    vendedor_nome,
                    loja_nome,
                    item.cliente_nome or "Não informado",
                    status,
                    valor,
                    motivo,
                    item.observacoes or ""
                ]
                ws.append(linha)

            # Estilizar as células de dados e aplicar formatação de moeda
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = font_dados
                    cell.border = border_fina
                    # Formatação específica para a coluna F (Valor Venda)
                    if cell.column == 6: 
                        cell.number_format = 'R$ #,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    # Centralizar Horário e Status
                    elif cell.column in [1, 5]: 
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="left")

            # Auto-ajuste de largura das colunas
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

            # Preparar a resposta http para download do arquivo Excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="relatorio_{nome_loja_slug}_{data_inicio}_a_{data_fim}.xlsx"'
            wb.save(response)
            return response

        return HttpResponse("Formato inválido. Use 'csv' ou 'xlsx'.", status=400)